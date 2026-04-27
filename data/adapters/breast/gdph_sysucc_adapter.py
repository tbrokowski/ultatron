"""
data/adapters/breast/gdph_sysucc_adapter.py  ·  GDPH & SYSUCC adapter
======================================================================
GDPH & SYSUCC (breast ultrasound classification): 2,405 images.
  Classes: benign, malignant
  Format:  PNG images, label encoded in filename prefix
  Source:  github.com / Alsharid 2025
  SonoDQS: bronze

Dataset layout on disk
-----------------------
GDPH_SYSUCC/
├── BIRADS&FOLD.xlsx     ← BIRADS scores + fold split assignments
├── GDPH/
│   ├── benign(1).png
│   ├── benign(2).png
│   ├── malignant(0).png
│   └── ...
└── SYSUCC/
    ├── benign(0).png
    ├── benign(1).png
    ├── malignant(0).png
    └── ...

Key observations
----------------
- Label comes from filename prefix: "benign(...)" or "malignant(...)".
- Two hospital sub-datasets: GDPH and SYSUCC — both iterated.
- BIRADS&FOLD.xlsx contains fold splits and BIRADS scores if available.
- No segmentation masks — classification only.
- sample_id includes sub-dataset name to avoid collisions between GDPH/SYSUCC.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Iterator, Optional

from data.adapters.base import BaseAdapter
from data.schema.manifest import USManifestEntry

# Matches "benign(1).png", "malignant(0).png"
_LABEL_RE = re.compile(r"^(benign|malignant)\((\d+)\)\.png$", re.IGNORECASE)

CLASSES = {
    "benign":    ("benign_lesion",    "breast_lesion_benign"),
    "malignant": ("malignant_lesion", "breast_lesion_malignant"),
}

SUB_DATASETS = ["GDPH", "SYSUCC"]


class GDPHSYSUCCAdapter(BaseAdapter):
    """
    Adapter for the GDPH & SYSUCC breast ultrasound dataset.

    Parameters
    ----------
    root : str | Path
        Root directory containing GDPH/, SYSUCC/, and BIRADS&FOLD.xlsx.
    split_override : str, optional
        If set, all entries get this split label.
    """

    DATASET_ID     = "GDPH-SYSUCC"
    ANATOMY_FAMILY = "breast"
    SONODQS        = "bronze"
    DOI            = "https://github.com/Alsharid2025"

    def __init__(
        self,
        root: str | Path,
        split_override: Optional[str] = None,
    ):
        super().__init__(root=root, split_override=split_override)
        self._fold_map = self._load_fold_map()

    # ── Private helpers ────────────────────────────────────────────────────

    def _load_fold_map(self) -> dict[str, str]:
        """
        Read BIRADS&FOLD.xlsx → mapping filename → split label.
        Returns empty dict if file is missing or openpyxl not available.
        """
        fold_map: dict[str, str] = {}
        xlsx_path = self.root / "BIRADS&FOLD.xlsx"
        if not xlsx_path.exists():
            return fold_map
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).lower() if c.value else "" for c in next(ws.iter_rows())]

            # Find filename and fold/split columns
            fname_col = next((i for i, h in enumerate(headers) if "file" in h or "name" in h or "image" in h), None)
            fold_col  = next((i for i, h in enumerate(headers) if "fold" in h or "split" in h), None)

            if fname_col is None or fold_col is None:
                return fold_map

            for row in ws.iter_rows(min_row=2, values_only=True):
                fname = str(row[fname_col]).strip() if row[fname_col] else ""
                fold  = str(row[fold_col]).strip()  if row[fold_col]  else ""
                if fname and fold:
                    # Map fold numbers to split labels
                    fold_map[fname] = "val" if fold in ("0", "test", "val") else "train"
        except Exception:
            pass
        return fold_map

    # ── Public interface ───────────────────────────────────────────────────

    def iter_entries(self) -> Iterator[USManifestEntry]:
        """Yield one USManifestEntry per image across GDPH and SYSUCC."""

        all_samples = []

        for sub in SUB_DATASETS:
            sub_dir = self.root / sub
            if not sub_dir.exists():
                continue
            for img_path in sorted(sub_dir.glob("*.png")):
                m = _LABEL_RE.match(img_path.name)
                if m:
                    cls_name = m.group(1).lower()
                    all_samples.append((img_path, cls_name, sub))

        n = len(all_samples)

        for i, (img_path, cls_name, sub) in enumerate(all_samples):
            label_raw, label_ontology = CLASSES.get(
                cls_name, ("unknown", "unknown")
            )

            # Split from xlsx, override, or inferred
            if self.split_override:
                split = self.split_override
            elif img_path.name in self._fold_map:
                split = self._fold_map[img_path.name]
            else:
                split = self._infer_split(img_path.stem, i, n)

            instance = self._make_instance(
                instance_id    = f"{sub}_{img_path.stem}",
                label_raw      = label_raw,
                label_ontology = label_ontology,
                mask_path      = None,
                is_promptable  = False,
            )

            yield self._make_entry(
                str(img_path),
                split,
                modality      = "image",
                instances     = [instance],
                has_mask      = False,
                task_type     = "classification",
                ssl_stream    = "image",
                is_promptable = False,
                probe_type    = "linear",
                source_meta   = {
                    "sub_dataset": sub,
                    "cls_name":    cls_name,
                    "doi":         self.DOI,
                },
            )
